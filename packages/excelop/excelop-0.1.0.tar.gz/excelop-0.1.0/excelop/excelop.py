import openpyxl
import os


class ExcelOp(object):  # Excel
    def __init__(self, file):
        self.file = file
        if os.path.exists(self.file):
            self.wb = openpyxl.load_workbook(self.file)
            sheets = self.wb.get_sheet_names()
            self.sheet = sheets[0]
            self.ws = self.wb[self.sheet]
        else:
            self.wb = openpyxl.Workbook()  # 创建一个工作薄
            self.ws = self.wb.active

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某行的值
    def set_row_value(self, row, list):
        for i in range(len(list)):
            self.set_cell_value(row, i+1, list[i])
    # 设置某列的值

    def set_col_value(self, col, list):
        for i in range(len(list)):
            self.set_cell_value(i+1, col, list[i])
    # 保存到文件

    def save(self, name=None):
        if name is None:
            name = self.file
        self.wb.save(name)

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row, colunm).value = cellvalue
        except:
            self.ws.cell(row, colunm).value = "*"
