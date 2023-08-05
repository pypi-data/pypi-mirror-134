import openpyxl as op
import six
import zipfile
from openpyxl.utils.cell import coordinate_to_tuple


class TableHeader(object):
    def __init__(self, title, width=None):
        self.title = title
        self.width = width


class WorkbookUnicodeReader(object):
    """Allows you to read data for an excel workbook.

     So far, it only allows to read data for workbooks that have one spreadsheet
    (spreadsheet selection is not available yet).

    Usage:

      with excel.WorkbookUnicodeReader(file name or stream) as reader:
        for row in reader:
          for cell in row:
            do_something(cell)

    cell in the example above will be a unicode string.
    """

    def __init__(self, file_input, strip_contents=False):
        if not zipfile.is_zipfile(file_input):
            raise IOError("Not an excel file")

        workbook = op.load_workbook(file_input)
        self.__active_sheet = workbook.active
        self.__strip_contents = strip_contents

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        pass

    def __iter__(self):
        def maybe_strip(value):
            return (
                value.strip()
                if isinstance(value, six.string_types) and self.__strip_contents
                else value
            )

        return six.moves.map(
            lambda row: [
                maybe_strip(six.text_type(cell.value)) if cell.value else ""
                for cell in row
            ],
            self.__active_sheet.rows,
        )


class WorkbookWriter(object):
    def __init__(self, output_stream):
        self.__out = output_stream
        self.__workbook = self._create_workbook()
        self.__spreadsheets = []
        self.__flushed = False

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.flush()

    def create_spreadsheet(self, title):
        if not self.__spreadsheets:
            self.__spreadsheets.append(self.__workbook.active)
        else:
            self.__spreadsheets.append(self.__workbook.create_sheet())

        self.__spreadsheets[-1].title = title
        return SpreadsheetWriter(self.__spreadsheets[-1])

    def _create_workbook(self):
        return op.Workbook()

    def flush(self):
        if not self.__flushed:
            self.__workbook.save(self.__out)
            self.__flushed = True


class SpreadsheetWriter(object):
    def __init__(self, sheet):
        self.__sheet = sheet

    def write_cell(self, rowcolumn, value, style=None):
        cell = self.__sheet.cell(rowcolumn)
        self._write_cell(cell, value, style)

    def _write_cell(self, cell, value, style=None, as_text=False):
        if not as_text:
            cell.value = value
        else:
            cell.set_explicit_value(value=value, data_type="s")

        if style:
            for attr, attrvalue in style.items():
                setattr(cell, attr, attrvalue)

    def write_table(
        self,
        topleft,
        headers,
        data,
        headerstyle=None,
        cellstyle=None,
        numerical_as_text=False,
    ):
        (row, column) = coordinate_to_tuple(topleft)

        active_cell = self.__sheet.cell(row, column)
        headerstyle = (
            headerstyle
            if headerstyle is not None
            else {"font": op.styles.Font(bold=True)}
        )

        for header_index, header in enumerate(headers):
            if header_index:
                active_cell = active_cell.offset(column=+1)
            self._write_cell(active_cell, header.title, headerstyle)

            if header.width is not None:
                self.__sheet.column_dimensions[
                    active_cell.column_letter
                ].width = header.width

        for row in data:
            active_cell = active_cell.offset(row=+1, column=-(len(headers) - 1))

            for column_index, column in enumerate(row):
                if column_index:
                    active_cell = active_cell.offset(column=+1)
                self._write_cell(
                    active_cell, column, cellstyle, as_text=numerical_as_text
                )


def interpret_sheet_as_table_with_headers(sheet):
    """Assumes sheet is a table and returns a list of dicts representing the data."""

    class Data(list):
        def __init__(self, title, headers):
            self.headers = headers
            self.title = title

    def maybe_strip(value):
        return value.strip() if isinstance(value, six.string_types) else value

    table_rows = sheet.iter_rows()
    headers = [maybe_strip(cell.value) for cell in next(table_rows)]

    if len(headers) != len(set(headers)):
        raise ValueError("Duplicate header columns found")

    if any(not header for header in headers):
        raise ValueError("Empty header cell found")

    data = Data(sheet.title, headers)

    for data_row in table_rows:
        data.append(
            {
                header: maybe_strip(value_cell.value)
                for header, value_cell in zip(headers, data_row)
            }
        )

    return data
