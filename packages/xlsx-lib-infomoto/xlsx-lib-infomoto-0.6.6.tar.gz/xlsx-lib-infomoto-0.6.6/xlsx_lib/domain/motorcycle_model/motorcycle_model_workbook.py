from io import BytesIO
from typing import Optional, List, Union

from openpyxl import \
    Workbook, \
    load_workbook

from xlsx_lib.domain.shared.NewImageFile import NewImageFile
from xlsx_lib.domain.motorcycle_model.SHEETNAMES_RELATIONSHIPS import SHEETNAMES_RELATIONSHIPS
from xlsx_lib.domain.smart_key.SmartKeyData import SmartKeyData
from xlsx_lib.domain.smart_key.SmartKeySheet import SmartKeySheet
from xlsx_lib.domain.abs.AbsData import AbsData
from xlsx_lib.domain.abs.AbsSheet import AbsSheet
from xlsx_lib.domain.autodiagnosis.AutodiagnosisData import AutodiagnosisData
from xlsx_lib.domain.autodiagnosis.AutodiagnosisSheet import AutodiagnosisSheet
from xlsx_lib.domain.distribution.distribution_sheet import DistributionData, DistributionSheet
from xlsx_lib.domain.power_supply.power_supply_component import PowerSupplyComponent
from xlsx_lib.domain.power_supply.power_supply_sheet import PowerSupplySheet
from xlsx_lib.domain.frame.frame_element import FrameElement
from xlsx_lib.domain.frame.frame_sheet import FrameSheet
from xlsx_lib.domain.motorcycle_model.motorcycle_model import MotorcycleModel
from xlsx_lib.domain.motorcycle_model.sheetnames import Sheetnames

from xlsx_lib.domain.engine.engine_section import EngineSection
from xlsx_lib.domain.engine.engine_sheet import EngineSheet

from xlsx_lib.domain.electronic.electronic_element import ElectronicElement
from xlsx_lib.domain.electronic.electronic_sheet import ElectronicSheet

from xlsx_lib.domain.generic_replacements.generic_replacement_sheet import GenericReplacementsSheet
from xlsx_lib.domain.generic_replacements.replacement import Replacement

from xlsx_lib.domain.tightening_specifications.tightening_specifications_sheet import TighteningSpecificationsSheet
from xlsx_lib.domain.tightening_specifications.specification_element import SpecificationElement


class MotorcycleModelWorkbook:
    def __init__(
            self,
            file: Union[BytesIO, str],
            filename: Optional[str],
    ):
        self.motorcycle_model: Optional[MotorcycleModel]
        self.distribution_images: Optional[List[NewImageFile]] = None
        self.smart_key_images: Optional[List[NewImageFile]] = None

        workbook: Workbook = load_workbook(filename=file)

        model_name: str = filename[filename.rfind("/") + 1:filename.rfind(".")] \
            .replace("FICHA ", "") \
            .strip() \
            .replace("  ", " ")

        generic_replacements: Optional[List[Replacement]] = None
        electronic_elements: Optional[List[ElectronicElement]] = None
        tightening_specifications: Optional[List[SpecificationElement]] = None
        engine_sections: Optional[List[EngineSection]] = None
        frame_elements: Optional[List[FrameElement]] = None
        power_supply_components: Optional[List[PowerSupplyComponent]] = None
        distribution: Optional[DistributionData] = None
        autodiagnosis: Optional[AutodiagnosisData] = None
        abs_data: Optional[AbsData] = None
        smart_key_data: Optional[SmartKeyData] = None

        for sheetname in [key for key in workbook.sheetnames
                          if key in SHEETNAMES_RELATIONSHIPS]:
            if Sheetnames.ENGINE == SHEETNAMES_RELATIONSHIPS[sheetname]:
                engine_sheet: EngineSheet = EngineSheet(worksheet=workbook[sheetname])
                engine_sections: List[EngineSection] = engine_sheet.engine_sections
                # TODO: Check if EngineSheet have distribution

            elif Sheetnames.GENERIC_REPLACEMENTS == SHEETNAMES_RELATIONSHIPS[sheetname]:
                generic_replacements = \
                    GenericReplacementsSheet(worksheet=workbook[sheetname]).get_generic_replacements()

            elif Sheetnames.ELECTRONIC == SHEETNAMES_RELATIONSHIPS[sheetname]:
                electronic_elements = \
                    ElectronicSheet(worksheet=workbook[sheetname]).get_electronic_elements()

            elif Sheetnames.TIGHTENING_TORQUES == SHEETNAMES_RELATIONSHIPS[sheetname]:
                tightening_specifications = \
                    TighteningSpecificationsSheet(worksheet=workbook[sheetname]).get_specification_elements()

            elif Sheetnames.FRAME == SHEETNAMES_RELATIONSHIPS[sheetname]:
                frame_elements = FrameSheet(worksheet=workbook[sheetname]).frame_elements

            elif Sheetnames.POWER_SUPPLY == SHEETNAMES_RELATIONSHIPS[sheetname]:
                power_supply_components = PowerSupplySheet(worksheet=workbook[sheetname]).components

            elif Sheetnames.DISTRIBUTION == SHEETNAMES_RELATIONSHIPS[sheetname]:
                distribution_sheet: DistributionSheet = DistributionSheet(worksheet=workbook[sheetname])

                distribution = distribution_sheet.distribution_data
                self.distribution_images = distribution_sheet.distribution_images

            elif Sheetnames.AUTODIAGNOSIS == SHEETNAMES_RELATIONSHIPS[sheetname]:
                autodiagnosis = AutodiagnosisSheet(worksheet=workbook[sheetname]).autodiagnosis

            elif Sheetnames.ABS == SHEETNAMES_RELATIONSHIPS[sheetname]:
                abs_data = AbsSheet(worksheet=workbook[sheetname]).abs_data

            elif Sheetnames.SMART_KEY == SHEETNAMES_RELATIONSHIPS[sheetname]:
                smart_key_sheet = SmartKeySheet(worksheet=workbook[sheetname])
                smart_key_data = smart_key_sheet.smart_key_data
                self.smart_key_images = smart_key_sheet.smart_key_images

        self.motorcycle_model = MotorcycleModel(
            model_name=model_name,
            generic_replacements=generic_replacements,
            tightening_specifications=tightening_specifications,
            electronic=electronic_elements,
            engine=engine_sections,
            frame=frame_elements,
            power_supply=power_supply_components,
            distribution=distribution,
            autodiagnosis=autodiagnosis,
            abs_data=abs_data,
            smart_key_data=smart_key_data
        )
